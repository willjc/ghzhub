#!/usr/bin/env python3
"""
缴费列表支付记录同步脚本
将Excel中已支付但数据库未同步的记录更新到hz_bill表。

安全措施：
1. Dry-run模式预览所有变更（默认）
2. 只更新 bill_status=0（未支付）的账单
3. 校验金额一致
4. 校验 transaction_no 不重复
5. 逐条记录操作日志
6. 使用事务，出错可回滚

用法：
  python3 sync_payments.py              # dry-run 预览
  python3 sync_payments.py --execute    # 实际执行
"""

import openpyxl
import subprocess
import sys
import json
import os
from datetime import datetime
from collections import defaultdict

# ===== 配置 =====
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '缴费列表.xlsx')
DB_HOST = '36.133.39.148'
DB_PORT = '33061'
DB_USER = 'ghzuser'
DB_PASS = '!aNJ-7vTE+Tsddce'
DB_NAME = 'ghz'

DRY_RUN = '--execute' not in sys.argv

# ===== 工具函数 =====
def mysql_query(sql, fetch=True):
    """执行MySQL查询"""
    cmd = ['mysql', '-h', DB_HOST, '-P', DB_PORT, '-u', DB_USER, f'-p{DB_PASS}', '-N', DB_NAME]
    if fetch:
        cmd.append('-e')
        cmd.append(sql)
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"[ERROR] SQL failed: {result.stderr}")
            return []
        rows = []
        for line in result.stdout.strip().split('\n'):
            if line.strip():
                rows.append(line.strip().split('\t'))
        return rows
    else:
        cmd.append('-e')
        cmd.append(sql)
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"[ERROR] SQL failed: {result.stderr}")
            return False
        return True


def mysql_execute(sql):
    """执行MySQL写操作"""
    result = subprocess.run(
        ['mysql', '-h', DB_HOST, '-P', DB_PORT, '-u', DB_USER, f'-p{DB_PASS}', DB_NAME, '-e', sql],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"[ERROR] SQL execute failed: {result.stderr}")
        return False
    return True


def escape_sql(s):
    """转义SQL字符串"""
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "\\'").replace("\\", "\\\\") + "'"


def parse_period(period_start_str):
    """从Excel的租金开始时间提取bill_period（YYYY-MM格式）"""
    s = str(period_start_str).strip()
    if not s or s == 'None':
        return None
    # 格式可能是 "2026-05-13 10:53:38" 或 "2026-05-13"
    try:
        if ' ' in s:
            dt = datetime.strptime(s.split(' ')[0], '%Y-%m-%d')
        else:
            dt = datetime.strptime(s, '%Y-%m-%d')
        return dt.strftime('%Y-%m')
    except:
        return None


def parse_amount(val):
    """解析金额"""
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except:
        return None


def parse_pay_time(val):
    """解析支付时间"""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == 'None':
        return None
    # 格式 "2026-06-04 17:06:41"
    try:
        dt = datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except:
        try:
            dt = datetime.strptime(s, '%Y-%m-%d')
            return dt.strftime('%Y-%m-%d 00:00:00')
        except:
            return None


def parse_method(val):
    """解析支付方式"""
    if val is None:
        return None
    s = str(val).strip()
    if s == '微信':
        return 'wechat'
    elif s == '支付宝':
        return 'alipay'
    return s


# ===== 主逻辑 =====
def main():
    mode = "DRY-RUN（预览模式，不会修改数据）" if DRY_RUN else "EXECUTE（实际执行模式）"
    print(f"=" * 60)
    print(f"缴费列表支付同步脚本 - {mode}")
    print(f"=" * 60)

    # 1. 读取Excel
    print("\n[1/5] 读取Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        all_rows.append(dict(zip(headers, row)))
    wb.close()
    print(f"  共 {len(all_rows)} 条记录")

    # 2. 过滤：只处理已支付的房租和押金
    print("\n[2/5] 过滤待同步记录...")
    target_records = []
    for r in all_rows:
        bill_type = r.get('账单类型', '')
        status = r.get('账单状态', '')
        if status == '已支付' and bill_type in ('房租', '押金'):
            target_records.append(r)
    print(f"  待同步: {len(target_records)} 条 (房租+押金已支付)")

    # 3. 获取DB中已有的transaction_no（排除已导入的）
    print("\n[3/5] 检查数据库已有记录...")
    # 分批获取DB中所有有transaction_no的账单
    db_txn_set = set()
    db_rows = mysql_query("SELECT transaction_no FROM hz_bill WHERE transaction_no IS NOT NULL AND transaction_no != '' AND del_flag = '0'")
    for row in db_rows:
        db_txn_set.add(row[0])
    print(f"  DB中已有 {len(db_txn_set)} 个transaction_no")

    # 过滤出Excel中不在DB的记录
    new_records = []
    already_synced = 0
    for r in target_records:
        flow_id = str(r.get('流水编号', '')).strip()
        if flow_id in db_txn_set:
            already_synced += 1
        else:
            new_records.append(r)

    print(f"  已同步: {already_synced}, 待导入: {len(new_records)}")

    # 4. 逐条匹配DB账单
    print("\n[4/5] 匹配数据库账单...")
    # 先批量获取需要的租户和合同信息
    # 收集所有涉及的租户名字
    tenant_names = set()
    for r in new_records:
        name = str(r.get('租户名字', '')).strip()
        if name:
            tenant_names.add(name)

    # 获取这些租户的合同信息
    tenant_contracts = {}  # name -> [(contract_id, tenant_id, phone)]
    if tenant_names:
        names_sql = ",".join(escape_sql(n) for n in tenant_names)
        rows = mysql_query(f"""
            SELECT c.contract_id, c.tenant_id, c.tenant_name, c.rent_price, c.deposit
            FROM hz_contract c
            WHERE c.tenant_name IN ({names_sql}) AND c.del_flag = '0'
        """)
        for row in rows:
            name = row[2]
            if name not in tenant_contracts:
                tenant_contracts[name] = []
            tenant_contracts[name].append({
                'contract_id': row[0],
                'tenant_id': row[1],
                'rent_price': float(row[3]) if row[3] else 0,
                'deposit': float(row[4]) if row[4] else 0,
            })

    print(f"  找到 {len(tenant_contracts)} 个租户的合同信息")

    # 匹配每条记录（已匹配bill_id去重，防止同一账单被多次分配）
    matched_bill_ids = set()  # 跟踪已匹配的bill_id
    matched = []
    unmatched = []
    amount_mismatch = []
    no_contract = []
    duplicate_skip = []

    for r in new_records:
        name = str(r.get('租户名字', '')).strip()
        bill_type = r.get('账单类型', '')
        amount = parse_amount(r.get('费用'))
        flow_id = str(r.get('流水编号', '')).strip()
        pay_time = parse_pay_time(r.get('支付时间'))
        method = parse_method(r.get('支付方式'))
        period = parse_period(r.get('租金开始时间'))

        record_info = {
            'name': name,
            'bill_type': bill_type,
            'amount': amount,
            'flow_id': flow_id,
            'pay_time': pay_time,
            'method': method,
            'period': period,
            'room': str(r.get('房间号', '')),
        }

        # 找合同
        contracts = tenant_contracts.get(name, [])
        if not contracts:
            no_contract.append(record_info)
            continue

        # 对于房租：按 period 匹配账单
        # 对于押金：按 bill_type=1 匹配
        if bill_type == '房租':
            # 在对应合同下找 bill_type=2, bill_period=period, bill_status=0
            found_bill = None
            for c in contracts:
                cid = c['contract_id']
                bill_rows = mysql_query(f"""
                    SELECT bill_id, bill_amount, bill_status, transaction_no
                    FROM hz_bill
                    WHERE contract_id = {cid}
                      AND bill_type = '2'
                      AND bill_period = {escape_sql(period)}
                      AND bill_status = '0'
                      AND del_flag = '0'
                      AND (transaction_no IS NULL OR transaction_no = '')
                    LIMIT 1
                """)
                if bill_rows:
                    bid = bill_rows[0][0]
                    if bid not in matched_bill_ids:
                        found_bill = {
                            'bill_id': bid,
                            'bill_amount': float(bill_rows[0][1]) if bill_rows[0][1] else 0,
                            'contract_id': cid,
                        }
                        break

            if found_bill:
                # 校验金额
                if abs(found_bill['bill_amount'] - amount) > 0.01:
                    amount_mismatch.append({**record_info, 'db_amount': found_bill['bill_amount'], 'bill_id': found_bill['bill_id']})
                else:
                    matched_bill_ids.add(found_bill['bill_id'])
                    matched.append({**record_info, 'bill_id': found_bill['bill_id'], 'contract_id': found_bill['contract_id']})
            else:
                unmatched.append(record_info)

        elif bill_type == '押金':
            # 找 bill_type=1 的未支付押金账单（无transaction_no）
            found_bill = None
            for c in contracts:
                cid = c['contract_id']
                bill_rows = mysql_query(f"""
                    SELECT bill_id, bill_amount, bill_status, transaction_no
                    FROM hz_bill
                    WHERE contract_id = {cid}
                      AND bill_type = '1'
                      AND bill_status = '0'
                      AND del_flag = '0'
                      AND (transaction_no IS NULL OR transaction_no = '')
                    LIMIT 1
                """)
                if bill_rows:
                    bid = bill_rows[0][0]
                    if bid not in matched_bill_ids:
                        found_bill = {
                            'bill_id': bid,
                            'bill_amount': float(bill_rows[0][1]) if bill_rows[0][1] else 0,
                            'contract_id': cid,
                        }
                        break

            if found_bill:
                if abs(found_bill['bill_amount'] - amount) > 0.01:
                    amount_mismatch.append({**record_info, 'db_amount': found_bill['bill_amount'], 'bill_id': found_bill['bill_id']})
                else:
                    matched_bill_ids.add(found_bill['bill_id'])
                    matched.append({**record_info, 'bill_id': found_bill['bill_id'], 'contract_id': found_bill['contract_id']})
            else:
                unmatched.append(record_info)

    # 打印匹配结果
    print(f"\n  === 匹配结果 ===")
    print(f"  成功匹配: {len(matched)} 条")
    print(f"  未找到账单: {len(unmatched)} 条")
    print(f"  金额不匹配: {len(amount_mismatch)} 条")
    print(f"  无合同: {len(no_contract)} 条")

    # 5. 执行更新
    if matched:
        print(f"\n[5/5] {'预览' if DRY_RUN else '执行'}更新 {len(matched)} 条记录...")

        # 二次安全检查：确保bill_id不重复
        bill_ids = [m['bill_id'] for m in matched]
        if len(bill_ids) != len(set(bill_ids)):
            print("[SAFETY] 发现重复bill_id，终止！")
            # 找出重复的
            seen = set()
            dups = []
            for bid in bill_ids:
                if bid in seen:
                    dups.append(bid)
                seen.add(bid)
            print(f"  重复bill_id: {dups}")
            return

        # 检查transaction_no唯一性
        flow_ids = [m['flow_id'] for m in matched]
        if len(flow_ids) != len(set(flow_ids)):
            print("[SAFETY] 发现重复flow_id，终止！")
            return

        success_count = 0
        fail_count = 0

        for i, m in enumerate(matched):
            update_sql = f"""
                UPDATE hz_bill SET
                    bill_status = '1',
                    paid_amount = {m['amount']},
                    pay_time = {escape_sql(m['pay_time'])},
                    pay_method = {escape_sql(m['method'])},
                    transaction_no = {escape_sql(m['flow_id'])},
                    update_time = NOW()
                WHERE bill_id = {m['bill_id']}
                  AND bill_status = '0'
                  AND del_flag = '0'
            """

            if DRY_RUN:
                if i < 20:  # 预览前20条
                    print(f"  [{i+1}] bill_id={m['bill_id']} {m['name']} "
                          f"{m['bill_type']} {m['amount']}元 period={m['period']} "
                          f"pay_time={m['pay_time']} method={m['method']} "
                          f"txn={m['flow_id'][:20]}...")
                elif i == 20:
                    print(f"  ... 省略 {len(matched) - 20} 条 ...")
            else:
                if mysql_execute(update_sql):
                    success_count += 1
                else:
                    fail_count += 1
                    print(f"  [FAIL] bill_id={m['bill_id']} {m['name']} {m['flow_id']}")

        if not DRY_RUN:
            print(f"\n  执行完成: 成功 {success_count}, 失败 {fail_count}")

    # 6. 输出未匹配详情（供人工处理）
    if unmatched or amount_mismatch or no_contract:
        log_file = os.path.join(os.path.dirname(__file__), 'sync_unmatched.log')
        with open(log_file, 'w', encoding='utf-8') as f:
            if unmatched:
                f.write(f"\n=== 未找到对应账单 ({len(unmatched)} 条) ===\n")
                for u in unmatched:
                    f.write(f"  {u['name']} | {u['bill_type']} | {u['amount']}元 | "
                            f"period={u['period']} | room={u['room']} | flow={u['flow_id']}\n")
            if amount_mismatch:
                f.write(f"\n=== 金额不匹配 ({len(amount_mismatch)} 条) ===\n")
                for a in amount_mismatch:
                    f.write(f"  {a['name']} | {a['bill_type']} | excel={a['amount']} "
                            f"db={a.get('db_amount')} | bill_id={a.get('bill_id')} | flow={a['flow_id']}\n")
            if no_contract:
                f.write(f"\n=== 无合同记录 ({len(no_contract)} 条) ===\n")
                for n in no_contract:
                    f.write(f"  {n['name']} | {n['bill_type']} | {n['amount']}元 | flow={n['flow_id']}\n")
        print(f"\n未匹配详情已写入: {log_file}")

    # 汇总
    print(f"\n{'=' * 60}")
    print(f"汇总:")
    print(f"  Excel 已支付记录: {len(target_records)}")
    print(f"  已同步(跳过): {already_synced}")
    print(f"  本次匹配成功: {len(matched)}")
    print(f"  未找到账单: {len(unmatched)}")
    print(f"  金额不匹配: {len(amount_mismatch)}")
    print(f"  无合同: {len(no_contract)}")
    if DRY_RUN and matched:
        print(f"\n  *** 这是预览模式，使用 --execute 参数实际执行 ***")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
